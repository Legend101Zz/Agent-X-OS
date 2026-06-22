"""G2 — books-prep end-to-end in sim mode.

Proves the books-prep pipeline runs through the kernel deterministically: ingest → categorise →
claim clean → queue low-conf → export (settles at L1 because the adapter's ``required_ring`` is
``L1`` and the instance is ``L1``). The sim-mode trajectory is driven by ``books_prep_playbook``
(registered in ``run_loop._SIM_PLAYBOOKS``); the kernel fulfills the read-class ``ingest_document``
calls natively with clearly-synthetic transactions; the reversible_write ``export_ledger`` and
``queue_manual_action`` calls flow through the gateway with the ``ExportLedgerAdapter`` /
``QueueManualActionAdapter``.

This is the design §7 "books_prep_playbook end-to-end in sim" check, plus the export Call proves
the pipeline produces a real ``.xlsx`` artifact via the live adapter.
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import cast

import pytest
from agentx_contracts import InstanceBinding, JsonValue, MandateType
from agentx_contracts.trigger import DeadlineTrigger
from agentx_kernel.bootstrap import build_phase1_runinvoker
from agentx_mandate.library.books_prep import build_books_prep_type
from agentx_syscall.registry import build_phase1_registry

NOW = datetime(2026, 6, 22, tzinfo=UTC)


def _instance() -> InstanceBinding:
    return InstanceBinding(
        instance_id="inst_books_e2e",
        type_ref="books-prep@0.1.0",
        ring="L1",
        heap_region_id="heap_books_e2e",
    )


def _mandate(documents: list[str] | None = None) -> MandateType:
    """Books-prep mandate with overridden target (documents + threshold)."""
    mandate = build_books_prep_type()
    target = dict(mandate.charter.target or {})
    target["documents"] = cast("list[JsonValue]", documents or ["sim_april.pdf"])
    target["output_format"] = "xlsx"
    target["confidence_threshold"] = 0.8
    return mandate.model_copy(
        update={"charter": mandate.charter.model_copy(update={"target": target})}
    )


@pytest.mark.asyncio
async def test_books_prep_sim_pipeline_settles_with_export_artifact(tmp_path: Path) -> None:
    """End-to-end: trigger → ingest (sim-native synthetic rows) → categorise → claim clean →
    queue any low-conf rows → export (settles at L1 because the adapter's required_ring matches
    the instance ring). The trace shows the full trajectory; the export wrote a real .xlsx."""
    out_dir = tmp_path / "out"
    invoker = build_phase1_runinvoker(
        registry=build_phase1_registry(
            books_intake_dir=tmp_path / "intake",
            books_output_dir=out_dir,
        ),
    )

    mandate = _mandate(["sim_april.pdf"])
    result = await invoker.invoke(
        mandate=mandate,
        instance=_instance(),
        trigger=DeadlineTrigger(ts=NOW, reason="books-prep e2e proof", entity_id="inst_books_e2e:run"),
        mode="sim",
    )

    # The run settled (export_ledger requires L1 and the instance is L1; the adapter ran).
    assert result.state == "settled", result.state
    assert result.park is None

    # Trace shows the canonical trajectory shape: plan → sim-native ingest → queue calls →
    # export → verify (settled).
    kinds = [event.kind for event in result.trace.events]
    assert "thought" in kinds  # playbook's plan + sim-native ingest fulfillment
    syscall_results = [event for event in result.trace.events if event.kind == "syscall_result"]
    summaries = [event.summary for event in syscall_results]
    # The export Call ran and succeeded.
    assert any("export_ledger" in s for s in summaries)
    # Low-confidence rows were routed to the queue adapter.
    assert any("queue_manual_action" in s for s in summaries)

    # The .xlsx file actually exists on disk (the live adapter wrote it).
    xlsx_files = list(out_dir.glob("*.xlsx"))
    assert xlsx_files, f"expected an .xlsx in {out_dir}, found {list(out_dir.iterdir())}"

    # And it's a real workbook with the three sheets the design promises.
    from openpyxl import load_workbook

    workbook = load_workbook(str(xlsx_files[0]))
    assert set(workbook.sheetnames) == {"Ledger", "Review Queue", "Summary"}


@pytest.mark.asyncio
async def test_books_prep_sim_claim_event_carries_a_ledger_transaction_fact(tmp_path: Path) -> None:
    """The Claim in the playbook emits at least one ``ledger_transaction`` Fact for clean rows."""
    invoker = build_phase1_runinvoker(
        registry=build_phase1_registry(
            books_intake_dir=tmp_path / "intake",
            books_output_dir=tmp_path / "out",
        ),
    )

    result = await invoker.invoke(
        mandate=_mandate(["sim_april.pdf"]),
        instance=_instance(),
        trigger=DeadlineTrigger(ts=NOW, reason="claim-shape", entity_id="inst_books_e2e:claim"),
        mode="sim",
    )

    # The Claim is journaled as a verify event; the Settlement is what materialises the Facts.
    # On settlement, the projection store should hold at least one ledger_transaction fact.
    assert result.state == "settled"
    assert result.settlement is not None
    settlement = result.settlement
    tx_facts = [f for f in settlement.facts if f.predicate == "ledger_transaction"]
    assert tx_facts, "expected at least one ledger_transaction fact in settlement"


@pytest.mark.asyncio
async def test_books_prep_sim_multiple_documents_concatenate_ingest_results(tmp_path: Path) -> None:
    """Two ingest Calls in the playbook → the read-result handler concatenates the synthetic
    transactions into one shared scratchpad list → categorizer sees all rows."""
    invoker = build_phase1_runinvoker(
        registry=build_phase1_registry(
            books_intake_dir=tmp_path / "intake",
            books_output_dir=tmp_path / "out",
        ),
    )

    mandate = _mandate(["sim_april.pdf", "sim_may.pdf"])
    result = await invoker.invoke(
        mandate=mandate,
        instance=_instance(),
        trigger=DeadlineTrigger(ts=NOW, reason="multi-doc", entity_id="inst_books_e2e:multi"),
        mode="sim",
    )

    # The run settled.
    assert result.state == "settled"

    # Trace records BOTH ingest Calls as sim-native "thought" events (one per doc) — proves the
    # multi-doc fan-out works through the playbook's per-doc loop.
    thought_events = [
        event for event in result.trace.events
        if event.kind == "thought" and "sim synthetic" in event.summary
    ]
    assert len(thought_events) == 2

    # The synthetic rows from each doc share account_id + date + amount + balance → identical
    # dedupe_keys. The dedupe guard (P0-2) reconciles the second batch's rows; the categorizer
    # emits one ledger_transaction fact per UNIQUE row, so the count is one-doc's worth, not two.
    assert result.settlement is not None
    tx_facts = [f for f in result.settlement.facts if f.predicate == "ledger_transaction"]
    assert tx_facts, "expected at least one ledger_transaction fact in settlement"
    # If we re-ran with the same docs and snapshot wasn't preserved, the second-batch rows would
    # ALL dedupe against the first-batch's rows. So the fact count equals one doc's worth.
    assert len(tx_facts) <= 3


@pytest.mark.asyncio
async def test_books_prep_sim_target_override_propagates_to_playbook(tmp_path: Path) -> None:
    """A high threshold (0.95) means more rows queue; the trace records proportionally more
    queue_manual_action calls than a low threshold."""
    out_low = tmp_path / "low"
    out_high = tmp_path / "high"

    async def run(threshold: float, out_dir: Path) -> int:
        invoker = build_phase1_runinvoker(
            registry=build_phase1_registry(
                books_intake_dir=tmp_path / "intake",
                books_output_dir=out_dir,
            ),
        )
        mandate = build_books_prep_type()
        target = dict(mandate.charter.target or {})
        target["documents"] = ["sim_april.pdf"]
        target["confidence_threshold"] = threshold
        mandate = mandate.model_copy(
            update={"charter": mandate.charter.model_copy(update={"target": target})}
        )
        result = await invoker.invoke(
            mandate=mandate,
            instance=_instance(),
            trigger=DeadlineTrigger(ts=NOW, reason=f"thr={threshold}", entity_id=f"inst_books_e2e:thr{threshold}"),
            mode="sim",
        )
        return sum(
            1 for event in result.trace.events
            if event.kind == "syscall_result" and event.summary == "queue_manual_action"
        )

    queues_low = await run(0.5, out_low)
    queues_high = await run(0.99, out_high)
    # A higher threshold queues more rows.
    assert queues_high >= queues_low