"""books-prep adapters: deterministic ingest (PDF/Excel/CSV) + ledger export.

PDF fixtures are generated in-process by a tiny, dependency-free PDF writer (``_make_pdf``) so the gate
needs no PDF-authoring library — the bytes are a real digital-text PDF that pdfplumber/pdfminer parse.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, cast

from agentx_contracts import JsonObject, SyscallRequest, SyscallResult
from agentx_syscall.books import ExportLedgerAdapter, IngestDocumentAdapter
from openpyxl import Workbook, load_workbook


def _txns(result: SyscallResult) -> list[dict[str, Any]]:
    return cast("list[dict[str, Any]]", result.output["transactions"])


def _req(name: str, args: dict[str, Any]) -> SyscallRequest:
    return SyscallRequest(
        name=name,
        args=cast(JsonObject, args),
        instance_id="inst_books",
        run_id="run_books_1",
        idempotency_key=f"run_books_1:{name}:1",
        ring="L1",
        risk_class="read" if name == "ingest_document" else "reversible_write",
    )


def _make_pdf(path: Path, lines: list[str], *, with_text: bool = True) -> None:
    """Write a minimal but valid single-page PDF; with_text=False yields a no-text ('scanned') page."""
    objects: list[bytes] = [
        b"<</Type/Catalog/Pages 2 0 R>>",
        b"<</Type/Pages/Kids[3 0 R]/Count 1>>",
        b"<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
        b"/Resources<</Font<</F1 4 0 R>>>>/Contents 5 0 R>>",
        b"<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
    ]
    if with_text:
        parts = [b"BT /F1 10 Tf 50 750 Td"]
        for index, line in enumerate(lines):
            esc = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)").encode("latin-1")
            parts.append((b" (" if index == 0 else b" 0 -14 Td (") + esc + b") Tj")
        parts.append(b" ET")
        stream = b"".join(parts)
    else:
        stream = b" "
    objects.append(b"<</Length " + str(len(stream)).encode() + b">>\nstream\n" + stream + b"\nendstream")

    out = b"%PDF-1.4\n"
    offsets: list[int] = []
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(out))
        out += str(index).encode() + b" 0 obj\n" + obj + b"\nendobj\n"
    xref_pos = len(out)
    size = len(objects) + 1
    out += b"xref\n0 " + str(size).encode() + b"\n0000000000 65535 f \n"
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (
        b"trailer\n<</Size " + str(size).encode() + b"/Root 1 0 R>>\nstartxref\n" + str(xref_pos).encode() + b"\n%%EOF"
    )
    path.write_bytes(out)


def _write_csv(path: Path, body: str) -> None:
    path.write_text(body, encoding="utf-8")


async def test_ingest_csv_extracts_rows_with_source_and_reconciled_confidence(tmp_path: Path) -> None:
    _write_csv(
        tmp_path / "stmt.csv",
        "Date,Narration,Debit,Credit,Balance,Ref\n"
        "02/04/2026,NEFT ACME TRADERS,,25000.00,125000.00,SIM1\n"
        "05/04/2026,UPI office supplies,1800.00,,123200.00,SIM2\n"
        "09/04/2026,NEFT GST challan,12000.00,,111200.00,SIM3\n",
    )
    adapter = IngestDocumentAdapter(intake_dir=tmp_path)
    result = await adapter.execute(_req("ingest_document", {"doc_id": "stmt.csv"}), None)

    assert result.status == "ok"
    txns = _txns(result)
    assert len(txns) == 3
    first, second, third = txns
    assert first["credit"] == 25000.0 and first["debit"] == 0.0
    assert second["debit"] == 1800.0
    # every row cites its source (doc id + page/line) and carries a dedupe key
    assert first["source"] == {"doc_id": "stmt.csv", "page": 1, "line": 2}
    assert str(first["dedupe_key"]).startswith("txn_")
    # reconciling rows score 1.0; the first (un-reconcilable) row scores 0.8 but is not suspect
    assert second["extraction_confidence"] == 1.0 and second["extraction_suspect"] is False
    assert third["extraction_confidence"] == 1.0
    assert first["extraction_confidence"] == 0.8


async def test_ingest_xlsx_extracts_rows(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Narration", "Debit", "Credit", "Balance", "Ref"])
    sheet.append(["02/04/2026", "NEFT ACME TRADERS", None, 25000.0, 125000.0, "X1"])
    sheet.append(["05/04/2026", "UPI supplies", 1800.0, None, 123200.0, "X2"])
    workbook.save(str(tmp_path / "stmt.xlsx"))

    adapter = IngestDocumentAdapter(intake_dir=tmp_path)
    result = await adapter.execute(_req("ingest_document", {"doc_id": "stmt.xlsx"}), None)

    assert result.status == "ok"
    txns = _txns(result)
    assert len(txns) == 2
    assert txns[1]["debit"] == 1800.0 and txns[1]["extraction_confidence"] == 1.0


async def test_ingest_digital_pdf_extracts_rows_and_resolves_debit_credit_from_balance(tmp_path: Path) -> None:
    _make_pdf(
        tmp_path / "stmt.pdf",
        [
            "Statement of Account XXXX1234",
            "02/04/2026 NEFT ACME TRADERS 25000.00 125000.00",
            "05/04/2026 UPI office supplies 1800.00 123200.00",
            "09/04/2026 NEFT GST challan 12000.00 111200.00",
        ],
    )
    adapter = IngestDocumentAdapter(intake_dir=tmp_path)
    result = await adapter.execute(_req("ingest_document", {"doc_id": "stmt.pdf"}), None)

    assert result.status == "ok", result.error
    txns = _txns(result)
    assert len(txns) == 3
    # debit vs credit resolved from the running-balance delta (PDF carries an unsigned amount)
    assert txns[0]["credit"] == 25000.0
    assert txns[1]["debit"] == 1800.0 and txns[1]["extraction_confidence"] == 1.0
    assert txns[2]["debit"] == 12000.0
    assert txns[0]["source"]["doc_id"] == "stmt.pdf"


async def test_ingest_scanned_pdf_returns_error_for_human_queue(tmp_path: Path) -> None:
    _make_pdf(tmp_path / "scan.pdf", [], with_text=False)
    adapter = IngestDocumentAdapter(intake_dir=tmp_path)
    result = await adapter.execute(_req("ingest_document", {"doc_id": "scan.pdf"}), None)

    assert result.status == "error"
    assert "scanned" in (result.error or "") or "no extractable" in (result.error or "")


async def test_ingest_missing_document_routes_to_error(tmp_path: Path) -> None:
    adapter = IngestDocumentAdapter(intake_dir=tmp_path)
    result = await adapter.execute(_req("ingest_document", {"doc_id": "nope.csv"}), None)
    assert result.status == "error"
    assert "not found" in (result.error or "")


async def test_ingest_structural_sanity_gate_bounces_a_mostly_broken_document(tmp_path: Path) -> None:
    # Balances that do not reconcile on most rows → whole-doc bounce (P0-3), not a confident wrong ledger.
    _write_csv(
        tmp_path / "broken.csv",
        "Date,Narration,Debit,Credit,Balance,Ref\n"
        "02/04/2026,row a,100.00,,500.00,A\n"
        "03/04/2026,row b,100.00,,999.00,B\n"  # expected 400, got 999 → suspect
        "04/04/2026,row c,100.00,,123.00,C\n"  # expected 899, got 123 → suspect
        "05/04/2026,row d,100.00,,7.00,D\n",  # expected 23, got 7 → suspect
    )
    adapter = IngestDocumentAdapter(intake_dir=tmp_path)
    result = await adapter.execute(_req("ingest_document", {"doc_id": "broken.csv"}), None)
    assert result.status == "error"
    assert "structural sanity" in (result.error or "")


async def test_export_ledger_writes_xlsx_with_ledger_review_and_summary_sheets(tmp_path: Path) -> None:
    rows: list[dict[str, object]] = [
        {
            "date": "02/04/2026",
            "narration": "NEFT ACME TRADERS",
            "debit": 0.0,
            "credit": 25000.0,
            "balance": 125000.0,
            "ref": "SIM1",
            "ledger_head": "Sales",
            "gst_treatment": "indeterminate_from_source",
            "confidence": 0.92,
            "vendor": "Acme Traders",
            "gstin": "27AABCU9603R1ZM",
            "state": "Maharashtra",
            "receivable_payable": "receivable",
            "missing_supporting_doc": False,
            "extraction_confidence": 0.8,
            "source": {"doc_id": "stmt.csv", "page": 1, "line": 2},
            "queued": False,
        },
        {
            "date": "05/04/2026",
            "narration": "UPI office supplies",
            "debit": 1800.0,
            "credit": 0.0,
            "balance": 123200.0,
            "ref": "SIM2",
            "ledger_head": "Office Expenses",
            "confidence": 0.4,
            "queued": True,
            "queue_reason": "low categorisation confidence",
            "source": {"doc_id": "stmt.csv", "page": 1, "line": 3},
        },
    ]
    adapter = ExportLedgerAdapter(output_dir=tmp_path)
    result = await adapter.execute(_req("export_ledger", {"filename": "april", "rows": rows}), None)

    assert result.status == "ok"
    assert result.output["row_count"] == 2 and result.output["queued_count"] == 1
    out_path = Path(str(result.output["path"]))
    assert out_path.exists() and out_path.name == "april.xlsx"

    workbook = load_workbook(str(out_path))
    assert set(workbook.sheetnames) == {"Ledger", "Review Queue", "Summary"}
    ledger = workbook["Ledger"]
    header = [cell.value for cell in ledger[1]]
    assert "Ledger Head" in header and "GST Treatment" in header and "Source Page/Line" in header
    assert ledger.max_row == 3  # header + 2 rows
    review = workbook["Review Queue"]
    assert review.max_row == 2  # header + 1 queued row


def test_adapters_have_well_formed_fixtures() -> None:
    assert IngestDocumentAdapter().fixtures[0].expect_status == "ok"
    assert ExportLedgerAdapter().fixtures[0].name == "export_ledger_smoke"
