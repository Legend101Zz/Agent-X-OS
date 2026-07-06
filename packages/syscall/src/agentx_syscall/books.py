"""books-prep syscall adapters — deterministic document parsing + ledger export (no LLM, no network).

The blueprint rule "no brain does I/O directly" puts ALL file reading/writing here, in deterministic
adapters, and leaves only judgment (categorisation) to the harness:

- ``IngestDocumentAdapter`` (``ingest_document``, READ) parses a digital-text PDF / Excel / CSV bank
  statement into structured transaction rows, each carrying a per-row **source citation** (doc id +
  page/line) and a deterministic **extraction_confidence** (does debit/credit reconcile against the
  row-to-row balance delta?). A scanned/image PDF (no extractable text) — or a document where too many
  rows fail structural checks — returns ``status="error"`` so the run routes it to the human queue
  rather than emitting a confident, wrong ledger (caveats P0-3).
- ``ExportLedgerAdapter`` (``export_ledger``, REVERSIBLE_WRITE, parks at L1) writes the categorised,
  source-cited rows to a clean ``.xlsx`` with a coverage/confidence summary sheet (caveats P2-2) and a
  CA review-queue sheet.

Neither adapter imports a credential root (invariant #2): parsing libs + ``agentx_contracts`` only.
"""

from __future__ import annotations

import csv
import hashlib
from collections.abc import Mapping
from datetime import datetime
from pathlib import Path
from typing import Any

from agentx_contracts import JsonObject, JsonValue, SyscallRequest, SyscallResult, SyscallTestCase
from agentx_contracts.security import Credential

from agentx_syscall.adapters import _AdapterBase, _error_result

# Canonical transaction shape (shared by the live adapter, the sim-native synthetic rows, the
# categoriser, and the exporter): every row carries date/narration/debit/credit/balance/ref, a
# source citation, deterministic extraction quality, and a heap dedupe key.
_RECONCILE_TOLERANCE = 0.01
_DEFAULT_STRUCTURAL_FAIL_THRESHOLD = 0.5

_DATE_HEADERS = ("date", "txn date", "transaction date", "value date", "tran date")
_NARRATION_HEADERS = ("narration", "description", "particulars", "details", "remarks", "transaction")
_DEBIT_HEADERS = ("debit", "withdrawal", "withdrawal amt", "dr", "paid in", "paid out", "amount debit")
_CREDIT_HEADERS = ("credit", "deposit", "deposit amt", "cr", "amount credit")
_BALANCE_HEADERS = ("balance", "closing balance", "running balance", "balance amt")
_REF_HEADERS = ("ref", "reference", "chq", "cheque", "chq no", "ref no", "instrument")


class _UnsupportedDocument(RuntimeError):
    """The document's extension is not a v0-supported format."""


class _UnrecognizedLayout(RuntimeError):
    """The document has digital text, but no line parsed as a transaction (unsupported layout)."""


# --- raw row extraction (per format) ----------------------------------------------------


def _num(value: object) -> float | None:
    """Parse a possibly-comma-grouped numeric cell; blank/dash → None; junk → None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "—", "."}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").lstrip("₹Rs ").strip()
    if text.endswith(("Dr", "DR", "Cr", "CR")):
        text = text[:-2].strip()
    try:
        amount = float(text)
    except ValueError:
        return None
    return -amount if negative else amount


def _match_header(headers: list[str], wanted: tuple[str, ...]) -> int | None:
    """Return the index of the first header cell that matches any wanted name (case-insensitive)."""
    lowered = [h.strip().lower() for h in headers]
    for index, head in enumerate(lowered):
        if head in wanted:
            return index
    for index, head in enumerate(lowered):
        if any(w in head for w in wanted):
            return index
    return None


def _row_from_columns(
    cells: list[str],
    cols: Mapping[str, int | None],
    *,
    page: int,
    line: int,
) -> dict[str, Any] | None:
    """Build a raw row from positional cells using a detected column map. None if there is no date."""

    def cell(key: str) -> str:
        index = cols.get(key)
        if index is None or index >= len(cells):
            return ""
        return str(cells[index]).strip()

    date = cell("date")
    if not date:
        return None
    debit = _num(cell("debit")) or 0.0
    credit = _num(cell("credit")) or 0.0
    balance = _num(cell("balance"))
    return {
        "date": date,
        "narration": cell("narration"),
        "debit": debit,
        "credit": credit,
        "balance": balance,
        "ref": cell("ref"),
        "page": page,
        "line": line,
    }


def _extract_tabular(rows: list[list[str]]) -> list[dict[str, Any]]:
    """Find the header row, map columns, and build raw rows (shared by CSV + XLSX paths)."""
    header_index = None
    cols: dict[str, int | None] = {}
    for index, cells in enumerate(rows[:25]):
        as_str = [str(c) for c in cells]
        date_col = _match_header(as_str, _DATE_HEADERS)
        if date_col is not None and (
            _match_header(as_str, _BALANCE_HEADERS) is not None
            or _match_header(as_str, _DEBIT_HEADERS) is not None
            or _match_header(as_str, _CREDIT_HEADERS) is not None
        ):
            header_index = index
            cols = {
                "date": date_col,
                "narration": _match_header(as_str, _NARRATION_HEADERS),
                "debit": _match_header(as_str, _DEBIT_HEADERS),
                "credit": _match_header(as_str, _CREDIT_HEADERS),
                "balance": _match_header(as_str, _BALANCE_HEADERS),
                "ref": _match_header(as_str, _REF_HEADERS),
            }
            break
    if header_index is None:
        return []
    out: list[dict[str, Any]] = []
    for line_offset, cells in enumerate(rows[header_index + 1 :], start=1):
        as_str = [str(c) if c is not None else "" for c in cells]
        if not any(c.strip() for c in as_str):
            continue
        row = _row_from_columns(as_str, cols, page=1, line=header_index + 1 + line_offset)
        if row is not None:
            out.append(row)
    return out


def _extract_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = [list(row) for row in csv.reader(handle)]
    return _extract_tabular(rows)


def _extract_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = [["" if cell is None else cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return _extract_tabular(rows)


def _extract_pdf(path: Path) -> list[dict[str, Any]]:
    """Extract transactions from a digital-text PDF.

    No extractable text at all → [] (scanned/image or empty). Text present but zero lines parse →
    ``_UnrecognizedLayout``, so the caller reports an unsupported layout instead of a scanned document.
    """
    import pdfplumber

    raw: list[dict[str, Any]] = []
    any_text = False
    with pdfplumber.open(str(path)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            if text.strip():
                any_text = True
            for line_number, text_line in enumerate(text.splitlines(), start=1):
                parsed = _parse_pdf_line(text_line, page=page_number, line_no=line_number)
                if parsed is not None:
                    raw.append(parsed)
    if not any_text:
        # pypdf fallback before declaring the document scanned/no-text.
        any_text = _pdf_has_text(path)
        if not any_text:
            return []
    if not raw:
        raise _UnrecognizedLayout(
            "digital text present but statement layout not recognized "
            "(unsupported date/column format); route to manual queue"
        )
    return raw


def _pdf_has_text(path: Path) -> bool:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    return any((page.extract_text() or "").strip() for page in reader.pages)


def _parse_pdf_line(text_line: str, *, page: int, line_no: int) -> dict[str, Any] | None:
    """Parse one statement line: a leading date, narration text, then trailing numeric tokens.

    Convention (digital statements): the last numeric token is the running balance and the
    second-to-last is the transaction amount; debit vs credit is resolved later from the balance delta.
    """
    tokens = text_line.split()
    if len(tokens) < 3:
        return None
    date, date_tokens = _leading_date(tokens)
    if date is None:
        return None
    numbers: list[tuple[int, float]] = []
    for index, token in enumerate(tokens):
        value = _num(token)
        if value is not None and any(ch.isdigit() for ch in token) and ("." in token or "," in token):
            numbers.append((index, value))
    if not numbers:
        return None
    balance = numbers[-1][1]
    amount = numbers[-2][1] if len(numbers) >= 2 else 0.0
    narration_end = numbers[0][0] if numbers else len(tokens)
    narration = " ".join(tokens[date_tokens:narration_end]).strip()
    return {
        "date": date,
        "narration": narration,
        "_pdf_amount": amount,
        "balance": balance,
        "ref": "",
        "page": page,
        "line": line_no,
        "debit": 0.0,
        "credit": 0.0,
    }


def _looks_like_date(token: str) -> bool:
    cleaned = token.replace("-", "/").replace(".", "/")
    parts = cleaned.split("/")
    return len(parts) == 3 and all(part.isdigit() for part in parts) and 1 <= len(parts[0]) <= 4


_MONTH_TOKENS = frozenset(
    {
        "jan", "january", "feb", "february", "mar", "march", "apr", "april", "may",
        "jun", "june", "jul", "july", "aug", "august", "sep", "sept", "september",
        "oct", "october", "nov", "november", "dec", "december",
    }
)


def _leading_date(tokens: list[str]) -> tuple[str | None, int]:
    """Return (date string, tokens consumed) for a leading date, or (None, 0).

    Accepts a single numeric token (``16/06/19``, ``16-06-2019``) or the spelled-month form
    ``16 Jun 19`` / ``16 June 2019`` spanning three tokens.
    """
    if _looks_like_date(tokens[0]):
        return tokens[0], 1
    if (
        len(tokens) >= 3
        and tokens[0].isdigit()
        and 1 <= len(tokens[0]) <= 2
        and tokens[1].lower() in _MONTH_TOKENS
        and tokens[2].isdigit()
        and len(tokens[2]) in (2, 4)
    ):
        return " ".join(tokens[:3]), 3
    return None, 0


# --- post-processing: source, extraction confidence, dedupe ------------------------------


def _statement_period(date: str) -> str:
    """Best-effort YYYY-MM for the (account, period) continuity scope; 'unknown' if undateable."""
    normalized = date.strip().replace(".", "/").replace("-", "/")
    for fmt in ("%d/%m/%Y", "%Y/%m/%d", "%d/%m/%y", "%m/%d/%Y", "%d %b %Y", "%d %b %y", "%d %B %Y", "%d %B %y"):
        try:
            return datetime.strptime(normalized, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return "unknown"


def _dedupe_key(account_id: str, row: Mapping[str, Any]) -> str:
    basis = "|".join(
        str(part)
        for part in (
            account_id,
            row.get("date", ""),
            row.get("debit", 0.0),
            row.get("credit", 0.0),
            row.get("balance", ""),
            row.get("ref", "") or row.get("narration", ""),
        )
    )
    return "txn_" + hashlib.sha1(basis.encode("utf-8")).hexdigest()[:16]


def _post_process(raw_rows: list[dict[str, Any]], doc_id: str) -> list[JsonObject]:
    """Assign source citations, resolve debit/credit from PDF balance deltas, score extraction quality."""
    account_id = f"acct:{doc_id}"
    period = _statement_period(raw_rows[0]["date"]) if raw_rows else "unknown"
    transactions: list[JsonObject] = []
    prev_balance: float | None = None
    for row in raw_rows:
        balance = row.get("balance")
        debit = float(row.get("debit") or 0.0)
        credit = float(row.get("credit") or 0.0)
        # PDF rows carry an unsigned amount; resolve debit vs credit from the balance delta.
        if "_pdf_amount" in row:
            amount = float(row["_pdf_amount"] or 0.0)
            if prev_balance is not None and isinstance(balance, (int, float)):
                if balance < prev_balance:
                    debit, credit = amount, 0.0
                else:
                    credit, debit = amount, 0.0
            else:
                credit, debit = amount, 0.0

        confidence, suspect = _score_extraction(prev_balance, debit, credit, balance)
        txn: JsonObject = {
            "date": str(row.get("date", "")),
            "narration": str(row.get("narration", "")),
            "debit": debit,
            "credit": credit,
            "balance": balance if isinstance(balance, (int, float)) else None,
            "ref": str(row.get("ref", "")),
            "source": {"doc_id": doc_id, "page": int(row.get("page", 1)), "line": int(row.get("line", 0))},
            "account_id": account_id,
            "statement_period": period,
            "extraction_confidence": confidence,
            "extraction_suspect": suspect,
        }
        txn["dedupe_key"] = _dedupe_key(account_id, txn)
        transactions.append(txn)
        if isinstance(balance, (int, float)):
            prev_balance = float(balance)
    return transactions


def _score_extraction(
    prev_balance: float | None, debit: float, credit: float, balance: object
) -> tuple[float, bool]:
    """Deterministic extraction quality: does debit/credit reconcile against the balance delta?

    - reconciles within tolerance → 1.0, not suspect
    - cannot reconcile (first row / missing balance) → 0.8, not suspect (unproven, not wrong)
    - fails reconciliation → 0.3, suspect (routes to the queue regardless of categorisation confidence)
    """
    if not isinstance(balance, (int, float)):
        return 0.6, True  # a row with no balance we could read is structurally suspect
    if prev_balance is None:
        return 0.8, False
    expected = prev_balance - debit + credit
    if abs(expected - float(balance)) <= _RECONCILE_TOLERANCE:
        return 1.0, False
    return 0.3, True


def _extract_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _extract_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_xlsx(path)
    if suffix == ".pdf":
        return _extract_pdf(path)
    raise _UnsupportedDocument(f"unsupported document type: {suffix or '(none)'}")


# --- the adapters -----------------------------------------------------------------------


class IngestDocumentAdapter(_AdapterBase):
    """Parse a digital-text PDF / Excel / CSV statement → structured, source-cited transaction rows."""

    def __init__(
        self,
        *,
        intake_dir: str | Path | None = None,
        structural_fail_threshold: float = _DEFAULT_STRUCTURAL_FAIL_THRESHOLD,
    ) -> None:
        self._intake_dir = Path(intake_dir) if intake_dir is not None else None
        self._structural_fail_threshold = structural_fail_threshold
        super().__init__(
            name="ingest_document",
            category="document",
            maturity_level=3,
            risk_class="read",
            required_ring="L0",
            tenant_auth="manual",
            fixtures=[
                SyscallTestCase(
                    name="ingest_document_smoke",
                    input={"doc_id": "sample_statement.csv"},
                    expect_status="ok",
                    expect_output_contains={"doc_id": "sample_statement.csv"},
                )
            ],
        )

    def _resolve_path(self, args: Mapping[str, Any]) -> Path | None:
        raw_path = args.get("path")
        if isinstance(raw_path, str) and raw_path:
            candidate = Path(raw_path)
            if candidate.is_absolute() or self._intake_dir is None:
                return candidate
            return self._intake_dir / raw_path
        doc_id = args.get("doc_id")
        if isinstance(doc_id, str) and doc_id and self._intake_dir is not None:
            return self._intake_dir / doc_id
        return None

    async def execute(self, req: SyscallRequest, cred: Credential | None) -> SyscallResult:
        path = self._resolve_path(req.args)
        doc_id = str(req.args.get("doc_id") or (path.name if path is not None else "document"))
        if path is None or not path.exists():
            return _error_result(
                req, self.name, self.maturity_level, f"document not found: {doc_id!r} (route to manual queue)"
            )
        try:
            raw_rows = _extract_rows(path)
        except _UnsupportedDocument as exc:
            return _error_result(req, self.name, self.maturity_level, str(exc))
        except _UnrecognizedLayout as exc:
            # Digital text exists but no line parsed → unsupported layout, still human queue (P0-3).
            return _error_result(req, self.name, self.maturity_level, str(exc))
        if not raw_rows:
            # No extractable transactions → scanned/image PDF or empty doc → human queue (P0-3 / non-goal).
            return _error_result(
                req,
                self.name,
                self.maturity_level,
                "no extractable transactions (scanned/image or empty document); route to manual queue",
            )
        transactions = _post_process(raw_rows, doc_id)
        suspect = sum(1 for txn in transactions if txn.get("extraction_suspect") is True)
        if transactions and suspect / len(transactions) > self._structural_fail_threshold:
            # Whole-document structural sanity gate: better to bounce the doc than hand the CA a
            # clean-looking lie (caveats P0-3).
            return _error_result(
                req,
                self.name,
                self.maturity_level,
                (
                    f"document failed structural sanity: {suspect}/{len(transactions)} rows suspect "
                    "(possible mis-parse); route to manual queue"
                ),
            )
        transactions_json: list[JsonValue] = list(transactions)
        return SyscallResult(
            status="ok",
            output={
                "transactions": transactions_json,
                "doc_id": doc_id,
                "row_count": len(transactions),
                "suspect_count": suspect,
                "unparsed": [],
            },
            idempotency_key=req.idempotency_key,
            fulfilled_by=self.name,
            maturity_used=self.maturity_level,
        )


_LEDGER_COLUMNS: list[tuple[str, str]] = [
    ("date", "Date"),
    ("narration", "Narration"),
    ("debit", "Debit"),
    ("credit", "Credit"),
    ("balance", "Balance"),
    ("ref", "Ref"),
    ("ledger_head", "Ledger Head"),
    ("gst_treatment", "GST Treatment"),
    ("confidence", "Confidence"),
    ("vendor", "Vendor"),
    ("gstin", "GSTIN"),
    ("state", "State"),
    ("receivable_payable", "Receivable/Payable"),
    ("missing_supporting_doc", "Missing Supporting Doc"),
    ("extraction_confidence", "Extraction Confidence"),
    ("source_doc", "Source Doc"),
    ("source_loc", "Source Page/Line"),
]


class ExportLedgerAdapter(_AdapterBase):
    """Write categorised, source-cited rows to a clean ``.xlsx`` ledger (parks at L1 for CA approval)."""

    def __init__(self, *, output_dir: str | Path | None = None) -> None:
        self._output_dir = Path(output_dir) if output_dir is not None else None
        super().__init__(
            name="export_ledger",
            category="document",
            maturity_level=3,
            risk_class="reversible_write",
            required_ring="L1",
            tenant_auth="manual",
            fixtures=[
                SyscallTestCase(
                    name="export_ledger_smoke",
                    input={"filename": "ledger.xlsx", "rows": []},
                    expect_status="ok",
                    expect_output_contains={"sheet": "Ledger"},
                )
            ],
        )

    async def execute(self, req: SyscallRequest, cred: Credential | None) -> SyscallResult:
        raw_rows = req.args.get("rows")
        rows = [row for row in raw_rows if isinstance(row, dict)] if isinstance(raw_rows, list) else []
        filename = str(req.args.get("filename") or f"ledger_{req.run_id}.xlsx")
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"
        out_dir = self._output_dir or Path.cwd()
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / filename
        queued = [row for row in rows if _is_queued(row)]
        try:
            self._write_workbook(out_path, rows, queued)
        except Exception as exc:  # noqa: BLE001 — surface any openpyxl failure as a clean syscall error
            return _error_result(req, self.name, self.maturity_level, f"failed to write ledger: {exc}")
        return SyscallResult(
            status="ok",
            output={
                "sheet": "Ledger",
                "path": str(out_path),
                "filename": filename,
                "row_count": len(rows),
                "queued_count": len(queued),
                "coverage": _coverage_summary(rows),
            },
            idempotency_key=req.idempotency_key,
            fulfilled_by=self.name,
            maturity_used=self.maturity_level,
        )

    def _write_workbook(self, out_path: Path, rows: list[JsonObject], queued: list[JsonObject]) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        ledger = workbook.active
        ledger.title = "Ledger"
        ledger.append([header for _, header in _LEDGER_COLUMNS])
        for row in rows:
            ledger.append([_cell_value(row, key) for key, _ in _LEDGER_COLUMNS])

        review = workbook.create_sheet("Review Queue")
        review.append(["Date", "Narration", "Amount", "Reason", "Source Doc", "Source Page/Line"])
        for row in queued:
            review.append(
                [
                    row.get("date", ""),
                    row.get("narration", ""),
                    _amount(row),
                    _queue_reason(row),
                    _source_doc(row),
                    _source_loc(row),
                ]
            )

        summary = workbook.create_sheet("Summary")
        summary.append(["Field", "Coverage", "Detail"])
        coverage = _coverage_summary(rows)
        summary.append(["transactions", coverage["transactions"], "rows in ledger"])
        summary.append(["queued_for_review", coverage["queued_for_review"], "low-confidence / extraction-suspect"])
        summary.append(["gstin_derivable", coverage["gstin_coverage"], "% of rows with a derived GSTIN"])
        summary.append(["vendor_resolved", coverage["vendor_coverage"], "% of rows with a resolved vendor"])
        summary.append(
            ["receivable_payable_tagged", coverage["receivable_payable_coverage"], "% of rows tagged"]
        )
        summary.append(["missing_supporting_doc", coverage["missing_supporting_doc_count"], "rows flagged"])
        workbook.save(str(out_path))


# --- export helpers ---------------------------------------------------------------------


def _is_queued(row: Mapping[str, Any]) -> bool:
    return bool(row.get("queued")) or bool(row.get("extraction_suspect"))


def _queue_reason(row: Mapping[str, Any]) -> str:
    reason = row.get("queue_reason")
    if isinstance(reason, str) and reason:
        return reason
    if row.get("extraction_suspect"):
        return "extraction_suspect: row failed balance reconciliation"
    return "low categorisation confidence"


def _amount(row: Mapping[str, Any]) -> float:
    debit = float(row.get("debit") or 0.0)
    credit = float(row.get("credit") or 0.0)
    return credit - debit


def _source_doc(row: Mapping[str, Any]) -> str:
    source = row.get("source")
    if isinstance(source, Mapping):
        return str(source.get("doc_id", ""))
    return ""


def _source_loc(row: Mapping[str, Any]) -> str:
    source = row.get("source")
    if isinstance(source, Mapping):
        return f"p{source.get('page', '')}/l{source.get('line', '')}"
    return ""


def _cell_value(row: Mapping[str, Any], key: str) -> JsonValue:
    if key == "source_doc":
        return _source_doc(row)
    if key == "source_loc":
        return _source_loc(row)
    value = row.get(key)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _coverage_summary(rows: list[JsonObject]) -> JsonObject:
    total = len(rows)

    def pct(predicate: Any) -> float:
        if total == 0:
            return 0.0
        return round(100.0 * sum(1 for row in rows if predicate(row)) / total, 1)

    return {
        "transactions": total,
        "queued_for_review": sum(1 for row in rows if _is_queued(row)),
        "gstin_coverage": pct(lambda row: bool(row.get("gstin"))),
        "vendor_coverage": pct(lambda row: bool(row.get("vendor"))),
        "receivable_payable_coverage": pct(lambda row: bool(row.get("receivable_payable"))),
        "missing_supporting_doc_count": sum(1 for row in rows if row.get("missing_supporting_doc")),
    }
